"""Report generation: a styled Excel workbook and a standalone HTML summary.
报告生成：带样式的 Excel 工作簿，以及一个独立的 HTML 总览页。

Excel: the full detail table (original data + QC columns) with risk levels
color-coded, plus a summary sheet. HTML: an at-a-glance dashboard with risk
distribution, rule hit rates (Chart.js), and which rules ran/were skipped.
Excel：完整明细表（原数据 + QC 列），风险等级按色标注，外加一个汇总表；
HTML：一页式看板——风险分布、各规则命中率（Chart.js）、以及哪些规则跑了/被跳过。
"""
from __future__ import annotations

import html
import json

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Risk-level fills for the Excel detail sheet. / Excel 明细表里各风险档的填充色。
_LEVEL_FILL = {
    "高": PatternFill(start_color="F8C9C4", end_color="F8C9C4", fill_type="solid"),
    "中": PatternFill(start_color="FCE8B2", end_color="FCE8B2", fill_type="solid"),
    "低": PatternFill(start_color="D7EAD9", end_color="D7EAD9", fill_type="solid"),
}
_QC_HEADER_FILL = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
_QC_HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_excel(detail: pd.DataFrame, summary: dict, path: str) -> str:
    """Write the detail + summary sheets, with QC columns styled. / 写出明细+汇总两张表，QC 列加样式。"""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="明细", index=False)
        _summary_frame(summary).to_excel(writer, sheet_name="汇总", index=False)

        ws = writer.sheets["明细"]
        cols = list(detail.columns)

        # Style QC column headers so they stand out from the user's columns.
        # 给 QC 列表头加样式，与用户原有列区分开。
        for j, col in enumerate(cols, start=1):
            if str(col).startswith("[QC] "):
                cell = ws.cell(row=1, column=j)
                cell.fill = _QC_HEADER_FILL
                cell.font = _QC_HEADER_FONT

        # Color each row's 风险等级 cell by level. / 按风险等级给每行的“风险等级”单元格上色。
        if "[QC] 风险等级" in cols:
            level_col = cols.index("[QC] 风险等级") + 1
            for i, level in enumerate(detail["[QC] 风险等级"].tolist(), start=2):
                fill = _LEVEL_FILL.get(str(level))
                if fill:
                    ws.cell(row=i, column=level_col).fill = fill

        # Reasonable column widths. / 设置合理的列宽。
        for j, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(j)].width = min(
                40, max(10, len(str(col)) + 4)
            )
    return path


def _summary_frame(summary: dict) -> pd.DataFrame:
    """Turn the summary dict into a two-column 指标/数值 table. / 把汇总字典转成“指标/数值”两列表。"""
    rows = [
        ("样本总数", summary.get("total", 0)),
        ("高风险（建议剔除）", summary.get("level_counts", {}).get("高", 0)),
        ("中风险（建议复核）", summary.get("level_counts", {}).get("中", 0)),
        ("低风险（保留）", summary.get("level_counts", {}).get("低", 0)),
        ("被标记总数", summary.get("flagged_total", 0)),
        ("是否启用 LLM", "是" if summary.get("llm_used") else "否"),
    ]
    for name, hits in summary.get("rule_hits", {}).items():
        rows.append((f"命中：{name}", hits))
    return pd.DataFrame(rows, columns=["指标", "数值"])


def render_html(summary: dict, title: str = "Cleanvey 质量检查报告") -> str:
    """A self-contained HTML dashboard (also reused by the web result page).
    一个自包含的 HTML 看板（网页结果页也复用它）。"""
    levels = summary.get("level_counts", {})
    hit_rates = summary.get("rule_hit_rates", {})
    skipped = summary.get("skipped_rules", [])

    rule_labels = list(hit_rates.keys())
    rule_values = [round(v * 100, 1) for v in hit_rates.values()]

    cards = "".join(
        _card(label, levels.get(key, 0), color)
        for label, key, color in [
            ("高风险 · 建议剔除", "高", "#c0392b"),
            ("中风险 · 建议复核", "中", "#d68910"),
            ("低风险 · 保留", "低", "#1e8449"),
        ]
    )

    skipped_html = ""
    if skipped:
        items = "".join(
            f"<li>{html.escape(s['rule'])} — {html.escape(s['reason'])}</li>" for s in skipped
        )
        skipped_html = f"<div class='note'><b>未运行的规则：</b><ul>{items}</ul></div>"

    return _HTML_TEMPLATE.format(
        title=html.escape(title),
        total=summary.get("total", 0),
        flagged=summary.get("flagged_total", 0),
        llm="已启用" if summary.get("llm_used") else "未启用",
        cards=cards,
        labels=json.dumps(rule_labels, ensure_ascii=False),
        values=json.dumps(rule_values),
        skipped=skipped_html,
    )


def write_html(summary: dict, path: str, title: str = "Cleanvey 质量检查报告") -> str:
    """Write the HTML dashboard to disk. / 把 HTML 看板写到磁盘。"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(render_html(summary, title))
    return path


def _card(label: str, value, color: str) -> str:
    """Render one summary card. / 渲染一张汇总卡片。"""
    return (
        f"<div class='card' style='border-top:4px solid {color}'>"
        f"<div class='num' style='color:{color}'>{value}</div>"
        f"<div class='lbl'>{html.escape(label)}</div></div>"
    )


_HTML_TEMPLATE = """<!doctype html>
<html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
  body{{font-family:-apple-system,'Segoe UI',Roboto,'PingFang SC',sans-serif;
       margin:0;background:#f5f6f8;color:#1f2933}}
  .wrap{{max-width:960px;margin:0 auto;padding:32px 20px}}
  h1{{font-size:22px;margin:0 0 4px}}
  .sub{{color:#6b7280;margin-bottom:24px}}
  .cards{{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:28px}}
  .card{{flex:1;min-width:160px;background:#fff;border-radius:10px;padding:20px;
         box-shadow:0 1px 3px rgba(0,0,0,.08)}}
  .num{{font-size:34px;font-weight:700}}
  .lbl{{color:#6b7280;font-size:14px;margin-top:4px}}
  .panel{{background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:20px}}
  .note{{background:#fff8e6;border:1px solid #fde68a;border-radius:8px;padding:12px 16px;font-size:14px}}
  .note ul{{margin:6px 0 0;padding-left:20px}}
</style></head>
<body><div class="wrap">
  <h1>{title}</h1>
  <div class="sub">共 {total} 份样本 · 被标记 {flagged} 份 · 语义检查（LLM）{llm}</div>
  <div class="cards">{cards}</div>
  <div class="panel"><h3>各规则命中率（%）</h3><canvas id="hits" height="120"></canvas></div>
  {skipped}
  <div class="sub">Generated by Cleanvey · synthetic/demo-safe</div>
</div>
<script>
new Chart(document.getElementById('hits'), {{
  type:'bar',
  data:{{labels:{labels},datasets:[{{label:'命中率 %',data:{values},
        backgroundColor:'#2563eb'}}]}},
  options:{{indexAxis:'y',plugins:{{legend:{{display:false}}}},
        scales:{{x:{{beginAtZero:true}}}}}}
}});
</script></body></html>"""
